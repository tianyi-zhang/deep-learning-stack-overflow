import os
from openpyxl import load_workbook
import re
from nltk.stem.wordnet import WordNetLemmatizer
from nltk.corpus import stopwords
from itertools import groupby

special_words = ['prosses', 'acsses']

class DataLoader():
    def __init__(self, wb, sheetname):
        self.sheetname = sheetname
        self.sheet = wb[self.sheetname]
        self.questions, self.categories, id_cate_dict = self._get_questions()

    def _get_questions(self):
        m_row = self.sheet.max_row
        id_cate_dict = []
        questions = []
        categories = []
        for i in range(2, m_row + 1):
            cell_obj = self.sheet.cell(row = i, column = 2)
            if not cell_obj.value:
                continue
            questions.append(cell_obj.value)
            categories.append(self.sheet.cell(row=i, column=6).value)
            try:
                id_cate_dict.append((self.sheet.cell(row=i, column=1).value.split('/')[-1].strip(), self.sheet.cell(row=i, column=6).value.strip().replace('\n', ' ')))
            except:
                print(i)
        return questions, categories, id_cate_dict

    def _check_pattern(self, token):
        method_patterns = [re.compile('\w+\.+'), re.compile('\w+\_+')]
        for method_pattern in method_patterns:
            if re.search(method_pattern, token):
                return '<method>'
            elif re.search('\_', token):
                return '<method>'
            else:
                return token

    def _process_ques(self):
        new_ques = []
        stop_words = set(stopwords.words('english'))
        for idx, line in enumerate(self.questions):
            if not line:
                continue
            terms = line.lower().split()
            for term_id, term in enumerate(terms):
                terms[term_id] = re.sub('[\"\(\)\`\=\:\[\]\-\,\?\+\/]*', ' ',term)
                if re.search(re.compile("(\w+)\'s"), terms[term_id]):
                    terms[term_id] = terms[term_id].split("'")[0]
                terms[term_id] = re.sub("[\']*", '', terms[term_id])
                terms[term_id] = re.sub(' +', '', terms[term_id])
                terms[term_id] = self._check_pattern(terms[term_id])
                terms[term_id] = WordNetLemmatizer().lemmatize(terms[term_id], 'v')
                if terms[term_id] not in special_words:
                    temp_token = WordNetLemmatizer().lemmatize(terms[term_id], 'n')
                    if terms[term_id] != temp_token and not re.search(r'ss$', terms[term_id]):
                        terms[term_id] = temp_token
                if terms[term_id] in stop_words or len(terms[term_id])==1:
                    terms[term_id] = ''
                if re.match(re.compile("(\d+)d$"), terms[term_id]):   # $ is the end of string
                    terms[term_id] = '<dim>'
                elif re.match(re.compile("(\d+)gb$"), terms[term_id]):   # $ is the end of string
                    terms[term_id] = '<mem>'
                elif terms[term_id].isdigit():
                    terms[term_id] = '<digit>'

            new_terms = [x[0] for x in groupby(terms)]
            new_terms = ' '.join(new_terms)
            new_terms = re.sub(r"[^0-9a-z,.?!<>' ]", ' ', new_terms)
            new_terms = re.sub(' +', ' ', new_terms)
            category = self.categories[idx].strip().replace('\n', ' ')
            new_ques.append(new_terms.strip()+'\t'+category+'\n')
        return new_ques

def save_ques(data, fn):
    fw = open(os.path.join('..', 'data', fn), 'w')
    for item in data:
        try:
            fw.write(item)
        except TypeError:
            print(item)
            continue
    fw.close()
    print('%s has finished writing...'%fn)


# wb = load_workbook(filename=os.path.join('..', 'data', 'dl_framework_stackoverflow_questions.xlsx'))
# tf_sheet = wb['TensorFlow(sample)']
# pt_sheet = wb['PyTorch(sample)']
# dj_sheet = wb['DeepLearning4J(sample)']
#
# tf_questions = DataLoader(wb, 'TensorFlow(sample)')._process_ques()
# pt_questions = DataLoader(wb, 'PyTorch(sample)')._process_ques()
# dj_questions = DataLoader(wb, 'DeepLearning4J(sample)')._process_ques()
#
# save_ques(tf_questions, 'tensforflow.txt')
# save_ques(dj_questions, 'dl4j.txt')
# save_ques(pt_questions, 'pytorch.txt')
# save_ques(tf_questions+pt_questions+dj_questions, 'total.txt')